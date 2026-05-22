from __future__ import annotations

import csv
import io
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from html import escape
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.db.models import Alert, ConfigEntry, MatchGroup, NormalizedRecord
from app.services.alert_engine import RULE_SHIP_AFTER_NO_FINANCE
from app.services.uninvoiced_dedupe import (
    dedupe_uninvoiced_entries,
    uninvoiced_amount_from_payload,
    viewer_display_order_no,
)
from app.services.normalize_service import parse_date
from app.services.uninvoiced_sorting_config import (
    DEFAULT_UNINVOICED_EXPORT_SORTING,
    UNINVOICED_EXPORT_SORTING_CONFIG_KEY,
)
from app.services.viewer_reminder_settings import load_disabled_customer_alert_keys, normalize_customer_key


_SH_TZ = ZoneInfo("Asia/Shanghai")
_HTML_TITLE = "超60天没开票"
_SHIP_DAYS = 60
_SORT_CONFIG_KEY = UNINVOICED_EXPORT_SORTING_CONFIG_KEY
_ORDER_SORT_FULLY_OUTBOUND_DATE_AMOUNT = "fully_outbound_then_order_date_desc_then_urgent_amount_desc"
_ORDER_SORT_FULLY_OUTBOUND_AMOUNT = "fully_outbound_then_urgent_amount_desc"
_CUSTOMER_SORT_AMOUNT_WITH_GROUPS = "amount_desc_with_sort_groups"
_CUSTOMER_SORT_GROUP_PRIORITY = "sort_group_priority_then_amount_desc"
_CUSTOMER_SORT_AMOUNT_ONLY = "amount_desc"


@dataclass
class _ExportLine:
    record_id: str
    source_row: int | None
    entry_line_no: str
    customer: str
    customer_order_no: str
    order_date: str
    item_name: str
    item_code: str
    product: str
    quantity: float | None
    amount: float | None
    order_total_amount: float | None
    tax_inclusive_unit_price: float | None
    executed_shipped_qty: float | None
    invoiced_qty: float | None
    uninvoiced_qty: float | None
    line_invoice_status: str
    order_unshipped_qty: float | None
    latest_outbound_date: str
    days_after_outbound: int | None
    days_over_threshold: int | None
    order_outbound_status: str
    line_outbound_status: str
    export_no: int | None = None


@dataclass
class _OrderSection:
    customer: str
    order_no: str
    order_outbound_status: str
    all_lines: list[_ExportLine] = field(default_factory=list)
    should_lines: list[_ExportLine] = field(default_factory=list)
    done_lines: list[_ExportLine] = field(default_factory=list)
    hold_lines: list[_ExportLine] = field(default_factory=list)

    @property
    def lines(self) -> list[_ExportLine]:
        return [*self.should_lines, *self.done_lines, *self.hold_lines]

    @property
    def total_amount(self) -> float:
        return _sum_known_amount(line.amount for line in self.lines)

    @property
    def total_amount_missing(self) -> bool:
        return any(line.amount is None for line in self.lines)

    @property
    def invoiced_amount(self) -> float:
        return _sum_known_amount(_line_invoiced_amount(line) for line in self.all_lines)

    @property
    def invoiced_amount_missing(self) -> bool:
        return self.source_total_amount_missing or any(_line_invoiced_amount(line) is None for line in self.all_lines)

    @property
    def urgent_amount(self) -> float:
        return _sum_known_amount(_line_uninvoiced_amount(line) for line in self.should_lines)

    @property
    def urgent_amount_missing(self) -> bool:
        return any(_line_uninvoiced_amount(line) is None for line in self.should_lines)

    @property
    def hold_amount(self) -> float:
        return _sum_known_amount(line.amount for line in self.hold_lines)

    @property
    def hold_amount_missing(self) -> bool:
        return any(line.amount is None for line in self.hold_lines)

    @property
    def source_total_amount(self) -> float:
        source_values = [
            line.order_total_amount
            for line in self.all_lines
            if not _is_missing_money_value(line.order_total_amount)
        ]
        if source_values:
            return round(float(source_values[0]), 6)
        return _sum_known_amount(
            line.amount for line in self.all_lines if not _is_missing_money_value(line.amount)
        )

    @property
    def source_total_amount_missing(self) -> bool:
        if any(not _is_missing_money_value(line.order_total_amount) for line in self.all_lines):
            return False
        return any(_is_missing_money_value(line.amount) for line in self.all_lines)

    @property
    def max_days_over_threshold(self) -> int | None:
        values = [line.days_over_threshold for line in self.should_lines if line.days_over_threshold is not None]
        return max(values) if values else None

    @property
    def max_days_after_outbound(self) -> int | None:
        values = [line.days_after_outbound for line in self.should_lines if line.days_after_outbound is not None]
        return max(values) if values else None

    @property
    def status_label(self) -> str:
        if _is_fully_outbound(self.order_outbound_status):
            return "整单已发完"
        return "整单未发完"

    @property
    def order_date_value(self) -> date | None:
        values = [parsed for line in self.all_lines if (parsed := parse_date(line.order_date)) is not None]
        return max(values) if values else None


@dataclass
class _CustomerSection:
    customer: str
    orders: list[_OrderSection] = field(default_factory=list)

    @property
    def urgent_amount(self) -> float:
        return _sum_known_amount(order.urgent_amount for order in self.orders)

    @property
    def urgent_amount_missing(self) -> bool:
        return any(order.urgent_amount_missing for order in self.orders)

    @property
    def hold_amount(self) -> float:
        return _sum_known_amount(order.hold_amount for order in self.orders)

    @property
    def hold_amount_missing(self) -> bool:
        return any(order.hold_amount_missing for order in self.orders)

    @property
    def max_days_over_threshold(self) -> int | None:
        values = [order.max_days_over_threshold for order in self.orders if order.max_days_over_threshold is not None]
        return max(values) if values else None

    @property
    def max_days_after_outbound(self) -> int | None:
        values = [order.max_days_after_outbound for order in self.orders if order.max_days_after_outbound is not None]
        return max(values) if values else None


def _sum_known_amount(values) -> float:
    total = 0.0
    for value in values:
        if value is not None:
            total += float(value)
    return round(total, 6)


def _line_invoiced_amount(line: _ExportLine) -> float | None:
    invoiced_qty = _effective_invoiced_qty(line)
    if invoiced_qty is None:
        return None
    if invoiced_qty <= 0:
        return 0.0
    return _line_amount_by_quantity(line, invoiced_qty)


def _line_uninvoiced_amount(line: _ExportLine) -> float | None:
    if line.uninvoiced_qty is None:
        return None
    if line.uninvoiced_qty <= 0:
        return 0.0
    return _line_amount_by_quantity(line, line.uninvoiced_qty)


def _line_amount_by_quantity(line: _ExportLine, quantity: float) -> float | None:
    if not _is_missing_money_value(line.tax_inclusive_unit_price):
        return round(float(line.tax_inclusive_unit_price) * float(quantity), 6)
    if _is_missing_money_value(line.amount) or line.quantity is None or line.quantity <= 0:
        return None
    return round(float(line.amount) * float(quantity) / float(line.quantity), 6)


def _is_missing_money_value(value: Any) -> bool:
    return value is None or _is_zero_number(value)


def _is_zero_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and abs(float(value)) < 1e-9


def _clean_text(value: Any, default: str = "") -> str:
    text = str(value or "").strip()
    return text or default


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _is_fully_outbound(value: Any) -> bool:
    return str(value or "").strip().lower() == "fully_outbound"


def _is_invoice_done(line: _ExportLine) -> bool:
    if line.uninvoiced_qty is not None and line.uninvoiced_qty <= 0:
        return True
    if line.quantity is not None and line.invoiced_qty is not None:
        return line.quantity > 0 and line.invoiced_qty >= line.quantity
    return False


def _effective_invoiced_qty(line: _ExportLine) -> float | None:
    if line.invoiced_qty is not None:
        return line.invoiced_qty
    if line.uninvoiced_qty is not None and line.uninvoiced_qty > 0:
        return 0.0
    return None


def _invoice_status_text(value: Any) -> str:
    text = str(value or "").strip()
    status_map = {
        "invoiced": "全部开票",
        "fully_invoiced": "全部开票",
        "全部开票": "全部开票",
        "partial": "部分开票",
        "partially_invoiced": "部分开票",
        "部分开票": "部分开票",
        "uninvoiced": "未开票",
        "not_invoiced": "未开票",
        "未开票": "未开票",
    }
    return status_map.get(text.lower(), status_map.get(text, text))


def _line_invoice_tag(line: _ExportLine) -> str:
    source_status = _invoice_status_text(line.line_invoice_status)
    if source_status == "全部开票":
        return "产品已开完"
    if source_status == "部分开票":
        return "产品已部分开票"
    if source_status == "未开票":
        return "产品未开票"
    invoiced_qty = _effective_invoiced_qty(line)
    if invoiced_qty is None or line.uninvoiced_qty is None:
        return "开票状态暂无"
    if line.uninvoiced_qty <= 0:
        return "产品已开完"
    if invoiced_qty <= 0:
        return "产品未开票"
    return "产品已部分开票"


def _line_invoice_status_text(line: _ExportLine) -> str:
    source_status = _invoice_status_text(line.line_invoice_status)
    if source_status:
        return source_status
    invoiced_qty = _effective_invoiced_qty(line)
    if invoiced_qty is None or line.uninvoiced_qty is None:
        return "暂无"
    if line.uninvoiced_qty <= 0:
        return "全部开票"
    if invoiced_qty <= 0:
        return "未开票"
    return "部分开票"


def _outbound_status_text(value: Any) -> str:
    text = str(value or "").strip()
    status_map = {
        "fully_outbound": "全部出库",
        "partially_outbound": "部分出库",
        "not_outbound": "未出库",
    }
    return status_map.get(text.lower(), text or "暂无")


def _fmt_num(value: float | None) -> str:
    if value is None:
        return "暂无"
    rounded = round(float(value), 6)
    if abs(rounded) < 1e-12:
        rounded = 0.0
    return f"{rounded:.6f}".rstrip("0").rstrip(".") or "0"


def _fmt_entry_line_no(value: Any) -> str:
    number = _to_float(value)
    if number is not None and number.is_integer():
        return str(int(number))
    return _clean_text(value)


def _normalize_quantity(value: float | None) -> float | int | None:
    if value is None:
        return None
    number = float(value)
    return int(number) if number.is_integer() else number


def _fmt_amount(value: float | None) -> str:
    if value is None:
        return "金额暂缺"
    return f"¥{float(value):,.2f}"


def _fmt_total_amount(value: float, has_missing: bool) -> str:
    if has_missing and _is_zero_number(value):
        return "金额暂缺"
    text = _fmt_amount(value)
    return f"{text}（部分金额暂缺）" if has_missing else text


def _fmt_days(value: int | None) -> str:
    return f"{value} 天" if value is not None else "暂无"


def _payload_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _core_from_record(record: NormalizedRecord) -> dict[str, Any]:
    payload = _payload_dict(record.payload_json)
    return _payload_dict(payload.get("core"))


def _customer_from_group(group: MatchGroup | None, payload: dict[str, Any]) -> str:
    if group and isinstance(group.summary_json, dict):
        aggregate = group.summary_json.get("aggregate")
        if isinstance(aggregate, dict):
            text = _clean_text(aggregate.get("customer"))
            if text:
                return text
    return _clean_text(payload.get("customer"), "未知客户")


def _order_no(value: Any) -> str:
    display = viewer_display_order_no(value)
    return _clean_text(display or value, "未知单据")


def _order_key(customer: Any, order_no: Any) -> tuple[str, str]:
    return (normalize_customer_key(_clean_text(customer)), _clean_text(order_no).lower())


def _product_label(item_name: Any, item_code: Any) -> str:
    name = _clean_text(item_name)
    code = _clean_text(item_code)
    if name and code:
        return f"{name}（{code}）"
    return name or code or "未知产品"


def _days_from_payload(payload: dict[str, Any]) -> int | None:
    value = _to_float(payload.get("days_after_outbound"))
    return int(value) if value is not None else None


def _line_from_record(
    record: NormalizedRecord,
    *,
    fallback_payload: dict[str, Any] | None = None,
    fallback_customer: str = "",
    generated_day: datetime | None = None,
) -> _ExportLine:
    payload = fallback_payload or {}
    core = _core_from_record(record)
    generated_day = generated_day or datetime.now(_SH_TZ)

    def pick(field: str) -> Any:
        value = core.get(field)
        if value is not None and value != "":
            return value
        return payload.get(field)

    customer = _clean_text(pick("customer") or fallback_customer, "未知客户")
    order_no = _order_no(pick("customer_order_no"))
    item_name = _clean_text(pick("item_name"))
    item_code = _clean_text(pick("item_code"))
    latest_outbound_date = _clean_text(pick("latest_outbound_date"))
    days_after_outbound = _days_from_payload(payload)
    if days_after_outbound is None and latest_outbound_date:
        outbound_day = parse_date(latest_outbound_date)
        if outbound_day:
            days_after_outbound = (generated_day.date() - outbound_day).days
    days_over_threshold = max(days_after_outbound - _SHIP_DAYS, 0) if days_after_outbound is not None else None

    return _ExportLine(
        record_id=_clean_text(record.id),
        source_row=record.source_row,
        entry_line_no=_fmt_entry_line_no(pick("entry_line_no")),
        customer=customer,
        customer_order_no=order_no,
        order_date=_clean_text(pick("biz_date")),
        item_name=item_name,
        item_code=item_code,
        product=_product_label(item_name, item_code),
        quantity=_to_float(pick("quantity")),
        amount=_to_float(pick("amount")),
        order_total_amount=_to_float(pick("order_total_amount")),
        tax_inclusive_unit_price=_to_float(pick("tax_inclusive_unit_price")),
        executed_shipped_qty=_to_float(pick("executed_shipped_qty")),
        invoiced_qty=_to_float(pick("invoiced_qty")),
        uninvoiced_qty=_to_float(pick("uninvoiced_qty")),
        line_invoice_status=_clean_text(pick("line_invoice_status")),
        order_unshipped_qty=_to_float(pick("order_unshipped_qty")),
        latest_outbound_date=latest_outbound_date or "暂无",
        days_after_outbound=days_after_outbound,
        days_over_threshold=days_over_threshold,
        order_outbound_status=_clean_text(pick("order_outbound_status")),
        line_outbound_status=_clean_text(pick("line_outbound_status")),
    )


def _line_from_alert_payload(entry: dict[str, Any], *, generated_day: datetime | None = None) -> _ExportLine:
    payload = _payload_dict(entry.get("payload"))
    generated_day = generated_day or datetime.now(_SH_TZ)
    customer = _clean_text(entry.get("customer"), "未知客户")
    order_no = _order_no(payload.get("customer_order_no"))
    item_name = _clean_text(payload.get("item_name"))
    item_code = _clean_text(payload.get("item_code"))
    latest_outbound_date = _clean_text(payload.get("latest_outbound_date"))
    days_after_outbound = _days_from_payload(payload)
    if days_after_outbound is None and latest_outbound_date:
        outbound_day = parse_date(latest_outbound_date)
        if outbound_day:
            days_after_outbound = (generated_day.date() - outbound_day).days
    days_over_threshold = max(days_after_outbound - _SHIP_DAYS, 0) if days_after_outbound is not None else None
    amount = uninvoiced_amount_from_payload(payload, _clean_text(entry.get("message")))

    return _ExportLine(
        record_id=_clean_text(payload.get("record_id") or f"alert:{getattr(entry.get('alert'), 'id', '')}"),
        source_row=int(payload.get("source_row") or 0) or None,
        entry_line_no=_fmt_entry_line_no(payload.get("entry_line_no")),
        customer=customer,
        customer_order_no=order_no,
        order_date=_clean_text(payload.get("biz_date")),
        item_name=item_name,
        item_code=item_code,
        product=_product_label(item_name, item_code),
        quantity=_to_float(payload.get("quantity")),
        amount=amount,
        order_total_amount=_to_float(payload.get("order_total_amount")),
        tax_inclusive_unit_price=_to_float(payload.get("tax_inclusive_unit_price")),
        executed_shipped_qty=_to_float(payload.get("executed_shipped_qty")),
        invoiced_qty=_to_float(payload.get("invoiced_qty")),
        uninvoiced_qty=_to_float(payload.get("uninvoiced_qty")),
        line_invoice_status=_clean_text(payload.get("line_invoice_status")),
        order_unshipped_qty=_to_float(payload.get("order_unshipped_qty")),
        latest_outbound_date=latest_outbound_date or "暂无",
        days_after_outbound=days_after_outbound,
        days_over_threshold=days_over_threshold,
        order_outbound_status=_clean_text(payload.get("order_outbound_status")),
        line_outbound_status=_clean_text(payload.get("line_outbound_status")),
    )


def _line_sort_key(line: _ExportLine) -> tuple[int, int, str]:
    entry_value = _to_float(line.entry_line_no)
    return (
        int(entry_value) if entry_value is not None else 999999,
        int(line.source_row or 999999),
        line.product,
    )


def _customer_priority_key(customer: _CustomerSection) -> tuple[float, str]:
    return (-customer.urgent_amount, customer.customer)


def _merge_dict_template(default: dict[str, Any], current: dict[str, Any] | None) -> dict[str, Any]:
    data = dict(default)
    current_data = dict(current) if isinstance(current, dict) else {}
    for key, value in current_data.items():
        if isinstance(value, dict) and isinstance(data.get(key), dict):
            data[key] = _merge_dict_template(data[key], value)
        else:
            data[key] = value
    return data


def _load_sorting_config(db: Session) -> dict[str, Any]:
    default = DEFAULT_UNINVOICED_EXPORT_SORTING
    item = db.get(ConfigEntry, _SORT_CONFIG_KEY)
    current = item.value_json if item and isinstance(item.value_json, dict) else {}
    return _merge_dict_template(default, current)


def _configured_sort_groups(config: dict[str, Any]) -> list[dict[str, Any]]:
    raw_groups = config.get("customer_sort_groups")
    if not isinstance(raw_groups, list):
        return []
    groups: list[dict[str, Any]] = []
    for index, raw_group in enumerate(raw_groups, start=1):
        if not isinstance(raw_group, dict):
            continue
        name = _clean_text(raw_group.get("name")) or f"排序组{index}"
        raw_keywords = raw_group.get("keywords")
        if isinstance(raw_keywords, str):
            keywords = [item.strip() for item in raw_keywords.split(",")]
        elif isinstance(raw_keywords, list):
            keywords = [_clean_text(item) for item in raw_keywords]
        else:
            keywords = []
        keywords = [keyword for keyword in keywords if keyword]
        if not keywords:
            continue
        priority = _to_float(raw_group.get("priority"))
        groups.append(
            {
                "name": name,
                "keywords": keywords,
                "priority": int(priority) if priority is not None else index,
            }
        )
    return sorted(groups, key=lambda group: (int(group["priority"]), str(group["name"])))


def _customer_sort_group(customer_name: str, config: dict[str, Any]) -> str:
    text = _clean_text(customer_name)
    for group in _configured_sort_groups(config):
        if any(keyword in text for keyword in group["keywords"]):
            return f"group:{group['priority']}:{group['name']}"
    return f"customer:{normalize_customer_key(text) or text}"


def _sort_customers(customers: list[_CustomerSection], config: dict[str, Any]) -> None:
    customer_sort = str(config.get("customer_sort") or _CUSTOMER_SORT_AMOUNT_WITH_GROUPS)
    if customer_sort == _CUSTOMER_SORT_AMOUNT_ONLY:
        customers.sort(key=_customer_priority_key)
        return

    group_priority: dict[str, tuple[float, str]] = {}
    for customer in customers:
        group = _customer_sort_group(customer.customer, config)
        priority = _customer_priority_key(customer)
        if group not in group_priority or priority < group_priority[group]:
            group_priority[group] = priority

    configured_groups = {
        f"group:{group['priority']}:{group['name']}": int(group["priority"])
        for group in _configured_sort_groups(config)
    }

    def group_sort_key(customer: _CustomerSection) -> tuple[Any, ...]:
        group = _customer_sort_group(customer.customer, config)
        if customer_sort == _CUSTOMER_SORT_GROUP_PRIORITY and group in configured_groups:
            return (0, configured_groups[group])
        if customer_sort == _CUSTOMER_SORT_GROUP_PRIORITY:
            return (1, *_customer_priority_key(customer))
        return group_priority[group]

    customers.sort(
        key=lambda customer: (
            group_sort_key(customer),
            _customer_priority_key(customer),
        )
    )


def _order_sort_key(order: _OrderSection, config: dict[str, Any]) -> tuple[Any, ...]:
    order_sort = str(config.get("order_sort") or _ORDER_SORT_FULLY_OUTBOUND_DATE_AMOUNT)
    fully_outbound_rank = 0 if _is_fully_outbound(order.order_outbound_status) else 1
    order_date_rank = -(order.order_date_value.toordinal() if order.order_date_value else 0)
    urgent_rank = -order.urgent_amount
    if order_sort == _ORDER_SORT_FULLY_OUTBOUND_AMOUNT:
        return (fully_outbound_rank, urgent_rank, order.order_no)
    if order_sort == _ORDER_SORT_FULLY_OUTBOUND_DATE_AMOUNT:
        return (fully_outbound_rank, order_date_rank, urgent_rank, order.order_no)
    return (fully_outbound_rank, order_date_rank, urgent_rank, order.order_no)


def _load_uninvoiced_entries(db: Session) -> list[dict[str, Any]]:
    rows = (
        db.query(Alert, MatchGroup)
        .join(MatchGroup, Alert.group_id == MatchGroup.id)
        .filter(Alert.alert_type == RULE_SHIP_AFTER_NO_FINANCE, Alert.status == "open")
        .all()
    )
    disabled_keys = load_disabled_customer_alert_keys(db)
    entries: list[dict[str, Any]] = []
    for alert, group in rows:
        payload = _payload_dict(alert.payload_json)
        customer = _customer_from_group(group, payload)
        customer_key = normalize_customer_key(customer)
        if customer_key and (customer_key, alert.alert_type) in disabled_keys:
            continue
        entries.append(
            {
                "alert": alert,
                "group": group,
                "payload": payload,
                "customer": customer,
                "message": alert.message,
                "created_at": alert.created_at,
                "last_changed_at": alert.created_at,
            }
        )
    return dedupe_uninvoiced_entries(entries)


def _load_current_order_records(db: Session) -> list[NormalizedRecord]:
    return (
        db.query(NormalizedRecord)
        .filter(
            NormalizedRecord.document_type == "order",
            NormalizedRecord.lifecycle_state == "active",
            NormalizedRecord.is_current_effective.is_(True),
        )
        .order_by(NormalizedRecord.source_row.asc(), NormalizedRecord.created_at.asc())
        .all()
    )


def _build_customer_sections(db: Session, generated_at: datetime) -> list[_CustomerSection]:
    sorting_config = _load_sorting_config(db)
    entries = _load_uninvoiced_entries(db)
    if not entries:
        return []

    records = _load_current_order_records(db)
    records_by_id = {record.id: record for record in records}
    records_by_order: dict[tuple[str, str], list[NormalizedRecord]] = defaultdict(list)
    for record in records:
        core = _core_from_record(record)
        customer = _clean_text(core.get("customer"))
        order_no = _order_no(core.get("customer_order_no"))
        if customer and order_no != "未知单据":
            records_by_order[_order_key(customer, order_no)].append(record)

    alert_lines_by_order: dict[tuple[str, str], list[_ExportLine]] = defaultdict(list)
    should_record_ids_by_order: dict[tuple[str, str], set[str]] = defaultdict(set)
    for entry in entries:
        payload = _payload_dict(entry.get("payload"))
        record = records_by_id.get(_clean_text(payload.get("record_id")))
        if record:
            line = _line_from_record(
                record,
                fallback_payload=payload,
                fallback_customer=_clean_text(entry.get("customer")),
                generated_day=generated_at,
            )
        else:
            line = _line_from_alert_payload(entry, generated_day=generated_at)
        key = _order_key(line.customer, line.customer_order_no)
        alert_lines_by_order[key].append(line)
        should_record_ids_by_order[key].add(line.record_id)

    customer_map: dict[str, _CustomerSection] = {}
    for key, alert_lines in alert_lines_by_order.items():
        context_records = records_by_order.get(key, [])
        context_lines = [
            _line_from_record(record, generated_day=generated_at)
            for record in context_records
        ]
        existing_ids = {line.record_id for line in context_lines}
        context_lines.extend(line for line in alert_lines if line.record_id not in existing_ids)
        should_ids = should_record_ids_by_order[key]

        order = _OrderSection(
            customer=alert_lines[0].customer,
            order_no=alert_lines[0].customer_order_no,
            order_outbound_status=alert_lines[0].order_outbound_status,
        )
        for line in sorted(context_lines, key=_line_sort_key):
            order.all_lines.append(line)
            if line.record_id in should_ids:
                order.should_lines.append(line)
                if not order.order_outbound_status:
                    order.order_outbound_status = line.order_outbound_status
            elif not _is_fully_outbound(line.line_outbound_status):
                order.hold_lines.append(line)
            elif _is_invoice_done(line):
                order.done_lines.append(line)

        if not order.should_lines:
            continue
        customer_key = normalize_customer_key(order.customer) or order.customer
        customer_section = customer_map.setdefault(customer_key, _CustomerSection(customer=order.customer))
        customer_section.orders.append(order)

    customers = list(customer_map.values())
    for customer in customers:
        customer.orders.sort(key=lambda order: _order_sort_key(order, sorting_config))
    _sort_customers(customers, sorting_config)
    return customers


def _assign_export_numbers(customers: list[_CustomerSection]) -> None:
    export_no = 1
    for customer in customers:
        for order in customer.orders:
            for line in order.all_lines:
                line.export_no = None
            for line in order.should_lines:
                line.export_no = export_no
                export_no += 1


def _render_line(line: _ExportLine) -> str:
    invoice_tag = _line_invoice_tag(line)
    uninvoiced_amount = _line_uninvoiced_amount(line)
    invoiced_qty = _effective_invoiced_qty(line)
    sequence_text = f"序号 {line.export_no}" if line.export_no is not None else "序号"
    metrics = [
        ("还差开票金额", _fmt_amount(uninvoiced_amount)),
        ("还差开票数量", _fmt_num(line.uninvoiced_qty)),
        ("已开票数量", _fmt_num(invoiced_qty)),
        ("订单数量", _fmt_num(line.quantity)),
        ("已出库数量", _fmt_num(line.executed_shipped_qty)),
        ("最近出库日期", line.latest_outbound_date),
        ("出库天数", _fmt_days(line.days_after_outbound)),
    ]
    metric_html = "".join(
        f"<span><b>{escape(label)}</b>{escape(str(value))}</span>" for label, value in metrics
    )
    return (
        f'<div class="product-line product-line--should">'
        f'<div class="product-main">'
        f'<span class="product-title">'
        f'<span class="line-seq">{escape(sequence_text)}</span>'
        f'<span class="product-name">{escape(line.product)}</span>'
        f"</span>"
        f'<span class="line-tag">{escape(invoice_tag)}</span>'
        f"</div>"
        f'<div class="line-metrics">{metric_html}</div>'
        f"</div>"
    )


def _render_order(order: _OrderSection, *, order_index: int, order_total: int) -> str:
    should_html = "".join(_render_line(line) for line in order.should_lines)
    order_index_text = f"订单 {order_index}/{order_total}"
    summary = [
        ("订单总金额", _fmt_total_amount(order.source_total_amount, order.source_total_amount_missing)),
        ("已开票金额", _fmt_total_amount(order.invoiced_amount, order.invoiced_amount_missing)),
        ("还差开票金额", _fmt_total_amount(order.urgent_amount, order.urgent_amount_missing)),
    ]
    summary_html = "".join(f"<span><b>{escape(label)}</b>{escape(value)}</span>" for label, value in summary)
    return f"""
      <article class="order-card">
        <header class="order-header">
          <div class="order-title">
            <strong>{escape(order.customer)}</strong>
            <span>{escape(order.order_no)}</span>
          </div>
          <span class="status-tag">{escape(order.status_label)}</span>
          <div class="amount-focus">
            <span>应催金额</span>
            <strong>{escape(_fmt_total_amount(order.urgent_amount, order.urgent_amount_missing))}</strong>
          </div>
        </header>
        <div class="order-summary">{summary_html}</div>
        {should_html}
        <div class="order-index">{escape(order_index_text)}</div>
      </article>
    """


def _customer_summary_html(customer: _CustomerSection, *, customer_index_text: str) -> str:
    summary = [
        ("应催金额", _fmt_total_amount(customer.urgent_amount, customer.urgent_amount_missing)),
        ("出库天数", _fmt_days(customer.max_days_after_outbound)),
        ("相关订单数", f"{len(customer.orders)} 笔"),
    ]
    summary_html = "".join(f"<span><b>{escape(label)}</b>{escape(value)}</span>" for label, value in summary)
    return f"""
      <div class="customer-summary">
        <header class="customer-heading">
          <h2>{escape(customer.customer)}</h2>
          <span class="customer-index">{escape(customer_index_text)}</span>
        </header>
        <div>{summary_html}</div>
      </div>
    """


def _render_html(customers: list[_CustomerSection], generated_at: datetime) -> str:
    total_urgent = _sum_known_amount(customer.urgent_amount for customer in customers)
    total_urgent_missing = any(customer.urgent_amount_missing for customer in customers)
    order_count = sum(len(customer.orders) for customer in customers)
    generated_text = generated_at.astimezone(_SH_TZ).strftime("%Y-%m-%d %H:%M")
    customer_summary_html: list[str] = []
    customer_detail_html: list[str] = []
    customer_total = len(customers)
    for customer_index, customer in enumerate(customers, start=1):
        customer_index_text = f"客户 {customer_index}/{customer_total}"
        customer_summary = _customer_summary_html(customer, customer_index_text=customer_index_text)
        customer_summary_html.append(customer_summary)
        order_total = len(customer.orders)
        orders_html = "".join(
            _render_order(order, order_index=index, order_total=order_total)
            for index, order in enumerate(customer.orders, start=1)
        )
        customer_detail_html.append(
            f"""
            <section class="customer-section">
              {customer_summary}
              <div class="order-grid">{orders_html}</div>
            </section>
            """
        )

    summary_body = "\n".join(customer_summary_html)
    detail_body = "\n".join(customer_detail_html)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{_HTML_TITLE}</title>
  <style>
    :root {{
      color: #161616;
      background: #ffffff;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; background: #f7f8f6; color: #161616; }}
    main {{ max-width: 1120px; margin: 0 auto; padding: 24px 16px 40px; }}
    .topbar {{ margin-bottom: 18px; border-bottom: 2px solid #222; padding-bottom: 14px; }}
    h1 {{ margin: 0 0 8px; font-size: 30px; font-weight: 750; letter-spacing: 0; }}
    .export-time {{ color: #555; font-size: 13px; }}
    .report-section {{ margin-top: 22px; }}
    .section-divider {{
      display: flex;
      align-items: center;
      gap: 12px;
      margin: 18px 0 14px;
      color: #161616;
      font-size: 18px;
      font-weight: 750;
    }}
    .section-divider::after {{
      content: "";
      flex: 1;
      height: 2px;
      background: #222;
    }}
    .overview {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 10px; margin: 18px 0 24px; }}
    .overview-item, .customer-summary, .order-card {{
      background: #fff;
      border: 1px solid #cfcfc8;
      border-radius: 8px;
      box-shadow: none;
    }}
    .overview-item {{ padding: 12px; }}
    .overview-item span, .amount-focus span {{ display: block; color: #666; font-size: 12px; }}
    .overview-item strong {{ display: block; margin-top: 4px; font-size: 20px; }}
    .summary-only-grid {{ display: grid; grid-template-columns: 1fr; gap: 12px; }}
    .customer-section {{ margin-top: 22px; }}
    .customer-summary {{ padding: 14px; margin-bottom: 10px; border-left: 4px solid #222; }}
    .customer-heading {{ display: flex; align-items: center; justify-content: space-between; gap: 10px; margin-bottom: 10px; }}
    .customer-heading h2 {{ flex: 1; min-width: 0; margin: 0; font-size: 21px; overflow-wrap: anywhere; }}
    .customer-index {{
      border: 1px solid #777;
      border-radius: 999px;
      padding: 3px 8px;
      color: #555;
      font-size: 12px;
      white-space: nowrap;
      background: #fff;
    }}
    .customer-summary div, .order-summary, .line-metrics {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 8px;
    }}
    .order-summary {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }}
    .customer-summary span, .order-summary span, .line-metrics span {{
      border-top: 1px solid #e3e3de;
      padding-top: 6px;
      min-width: 0;
      font-size: 13px;
    }}
    .customer-summary b, .order-summary b, .line-metrics b {{
      display: block;
      color: #666;
      font-size: 11px;
      font-weight: 600;
      margin-bottom: 3px;
    }}
    .order-grid {{ display: grid; grid-template-columns: 1fr; gap: 12px; }}
    .order-card {{ padding: 14px; break-inside: avoid; page-break-inside: avoid; }}
    .order-header {{ display: grid; grid-template-columns: 1fr auto; gap: 10px; align-items: start; border-bottom: 1px solid #d8d8d2; padding-bottom: 10px; }}
    .order-title strong, .order-title span {{ display: block; }}
    .order-title strong {{ font-size: 17px; }}
    .order-title span {{ margin-top: 3px; font-size: 13px; color: #555; }}
    .status-tag, .line-tag {{
      border: 1px solid #777;
      border-radius: 999px;
      padding: 3px 8px;
      font-size: 12px;
      white-space: nowrap;
      background: #fff;
    }}
    .amount-focus {{ grid-column: 1 / -1; padding-top: 6px; }}
    .amount-focus strong {{ display: block; margin-top: 2px; font-size: 24px; }}
    .order-summary {{ margin: 10px 0 12px; }}
    .order-index {{ margin-top: 10px; text-align: right; color: #666; font-size: 12px; }}
    h4 {{ margin: 14px 0 8px; font-size: 14px; border-bottom: 1px solid #ecece7; padding-bottom: 5px; }}
    .product-line {{ border: 1px solid #dadad4; border-radius: 7px; padding: 10px; margin-top: 8px; background: #fff; }}
    .product-main {{ display: grid; grid-template-columns: 1fr auto; gap: 8px; align-items: center; }}
    .product-title {{ display: flex; flex-wrap: wrap; gap: 8px; align-items: center; min-width: 0; }}
    .line-seq {{ border: 1px solid #777; border-radius: 999px; padding: 2px 7px; background: #f7f7f3; color: #555; font-size: 12px; white-space: nowrap; }}
    .product-name {{ font-weight: 700; min-width: 0; overflow-wrap: anywhere; }}
    .line-metrics {{ margin-top: 8px; grid-template-columns: repeat(2, minmax(0, 1fr)); }}
    @media (max-width: 640px) {{
      main {{ padding: 18px 12px 32px; }}
      .overview, .customer-summary div {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .order-summary {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }}
      .order-grid {{ grid-template-columns: 1fr; }}
      h1 {{ font-size: 26px; }}
    }}
    @media print {{
      @page {{ size: A4; margin: 10mm; }}
      body {{ background: #fff; }}
      main {{ max-width: none; padding: 0; }}
      .overview {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }}
      .order-grid {{ grid-template-columns: 1fr; gap: 6mm; }}
      .order-card, .customer-summary, .overview-item {{ box-shadow: none; }}
      .customer-section {{ break-inside: auto; }}
      .detail-report {{ break-before: page; page-break-before: always; }}
      .section-divider {{ margin-top: 0; }}
    }}
  </style>
</head>
<body>
  <main>
    <header class="topbar">
      <h1>{_HTML_TITLE}</h1>
      <div class="export-time">导出时间：{escape(generated_text)}</div>
    </header>
    <section class="report-section summary-report" aria-label="汇总">
      <div class="section-divider">汇总</div>
      <section class="overview" aria-label="大汇总">
        <div class="overview-item"><span>客户数</span><strong>{len(customers)}</strong></div>
        <div class="overview-item"><span>相关订单数</span><strong>{order_count}</strong></div>
        <div class="overview-item"><span>应催金额合计</span><strong>{escape(_fmt_total_amount(total_urgent, total_urgent_missing))}</strong></div>
      </section>
      <div class="summary-only-grid" aria-label="客户小汇总">
        {summary_body}
      </div>
    </section>
    <section class="report-section detail-report" aria-label="明细">
      <div class="section-divider">明细</div>
      {detail_body}
    </section>
  </main>
</body>
</html>
"""


def export_uninvoiced_html(db: Session, generated_at: datetime | None = None) -> bytes | None:
    generated_at = generated_at or datetime.now(timezone.utc).astimezone(_SH_TZ)
    customers = _build_customer_sections(db, generated_at)
    if not customers:
        return None
    _assign_export_numbers(customers)
    return _render_html(customers, generated_at).encode("utf-8")


_SUMMARY_HEADERS = ["序号", "客户", "应催金额", "应催订单数", "应催产品行数", "最长出库天数"]
_DETAIL_HEADERS = [
    "序号",
    "核对状态",
    "客户",
    "单据日期",
    "单据编号",
    "分录行号",
    "商品名称",
    "商品编码",
    "已开票金额",
    "未开票金额",
    "价税合计",
    "成交金额",
    "含税单价",
    "数量",
    "行出库状态",
    "行已执行已出库数量",
    "行开票状态",
    "行已开票数量",
    "行未开票数量",
    "最近出库日期",
    "已超天数",
]

_DETAIL_STATUS_FILLS = {
    "应催": PatternFill("solid", fgColor="FCE4D6"),
    "未发完暂不催": PatternFill("solid", fgColor="FFF2CC"),
    "已开完不催": PatternFill("solid", fgColor="E2F0D9"),
    "未满60天暂不催": PatternFill("solid", fgColor="DDEBF7"),
    "暂不催": PatternFill("solid", fgColor="E7E6E6"),
}
_STATUS_EXPLANATIONS = [
    ("应催", "这条产品行已经出库满60天，并且还有未开票数量，需要催票。"),
    ("未发完暂不催", "这条产品行还没有全部出库，先不催票，但因为同订单出现在 HTML 里，所以放进 Excel 方便核对。"),
    ("已开完不催", "这条产品行已经开票完成，不需要催票。"),
    ("未满60天暂不催", "这条产品行已经出库但未满60天，暂时不催票。"),
    ("暂不催", "不属于上面几类的同订单产品行，放进 Excel 作为核对补充。"),
]


def _customer_line_count(customer: _CustomerSection) -> int:
    return sum(len(order.should_lines) for order in customer.orders)


def _detail_line_status(order: _OrderSection, line: _ExportLine) -> str:
    should_ids = {item.record_id for item in order.should_lines}
    if line.record_id in should_ids:
        return "应催"
    if not _is_fully_outbound(line.line_outbound_status):
        return "未发完暂不催"
    if _is_invoice_done(line):
        return "已开完不催"
    if line.days_after_outbound is not None and line.days_after_outbound < _SHIP_DAYS:
        return "未满60天暂不催"
    return "暂不催"


def _detail_rows(customers: list[_CustomerSection]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for customer in customers:
        for order in customer.orders:
            for line in order.all_lines:
                rows.append(
                    [
                        line.export_no,
                        _detail_line_status(order, line),
                        customer.customer,
                        line.order_date,
                        order.order_no,
                        line.entry_line_no,
                        line.item_name,
                        line.item_code,
                        _line_invoiced_amount(line),
                        _line_uninvoiced_amount(line),
                        line.amount,
                        order.source_total_amount,
                        line.tax_inclusive_unit_price,
                        _normalize_quantity(line.quantity),
                        _outbound_status_text(line.line_outbound_status),
                        _normalize_quantity(line.executed_shipped_qty),
                        _line_invoice_status_text(line),
                        _normalize_quantity(line.invoiced_qty),
                        _normalize_quantity(line.uninvoiced_qty),
                        "" if line.latest_outbound_date == "暂无" else line.latest_outbound_date,
                        line.days_over_threshold,
                    ]
                )
    return rows


def _summary_rows(customers: list[_CustomerSection]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for index, customer in enumerate(customers, start=1):
        rows.append(
            [
                index,
                customer.customer,
                customer.urgent_amount,
                len(customer.orders),
                _customer_line_count(customer),
                customer.max_days_after_outbound,
            ]
        )
    return rows


def _style_sheet(
    ws,
    *,
    widths: list[int],
    amount_columns: set[int],
    number_columns: set[int],
    unit_price_columns: set[int] | None = None,
    status_column: int | None = None,
    order_column: int | None = None,
) -> None:
    unit_price_columns = unit_price_columns or set()
    header_fill = PatternFill("solid", fgColor="E8EFE8")
    border_side = Side(style="thin", color="808080")
    header_side = Side(style="medium", color="666666")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    header_border = Border(left=border_side, right=border_side, top=header_side, bottom=header_side)
    group_top_border = Border(left=border_side, right=border_side, top=header_side, bottom=border_side)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = True
    ws.print_options.gridLines = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45

    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    previous_order = None
    for row in ws.iter_rows():
        current_order = None
        if order_column is not None and row[0].row > 1:
            current_order = ws.cell(row=row[0].row, column=order_column).value
        starts_new_order = bool(order_column and row[0].row > 2 and current_order != previous_order)
        row_fill = None
        if status_column is not None and row[0].row > 1:
            row_fill = _DETAIL_STATUS_FILLS.get(str(ws.cell(row=row[0].row, column=status_column).value or ""))
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = group_top_border if starts_new_order else border
            if cell.row > 1 and _is_zero_number(cell.value):
                cell.value = None
            if cell.row == 1:
                cell.font = Font(bold=True, color="1F2A23")
                cell.fill = header_fill
                cell.border = header_border
            elif cell.column in amount_columns:
                cell.number_format = '#,##0.00'
            elif cell.column in unit_price_columns:
                cell.number_format = '#,##0.#####'
            elif cell.column in number_columns:
                cell.number_format = '#,##0'
            if row_fill is not None and cell.row > 1:
                cell.fill = row_fill
        if order_column is not None and row[0].row > 1:
            previous_order = current_order
        ws.row_dimensions[row[0].row].height = 24 if row[0].row == 1 else 34


def _build_status_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("颜色说明")
    sheet.append(["颜色", "核对状态", "说明"])
    for status, description in _STATUS_EXPLANATIONS:
        sheet.append(["", status, description])

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = True
    sheet.print_options.gridLines = True
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 68

    header_fill = PatternFill("solid", fgColor="E8EFE8")
    border_side = Side(style="thin", color="808080")
    header_side = Side(style="medium", color="666666")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    header_border = Border(left=border_side, right=border_side, top=header_side, bottom=header_side)
    for row in sheet.iter_rows():
        status = str(sheet.cell(row=row[0].row, column=2).value or "")
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True, color="1F2A23")
                cell.fill = header_fill
                cell.border = header_border
            elif status in _DETAIL_STATUS_FILLS:
                cell.fill = _DETAIL_STATUS_FILLS[status]
        sheet.row_dimensions[row[0].row].height = 24 if row[0].row == 1 else 40


def export_uninvoiced_excel(db: Session, generated_at: datetime | None = None) -> bytes | None:
    generated_at = generated_at or datetime.now(timezone.utc).astimezone(_SH_TZ)
    customers = _build_customer_sections(db, generated_at)
    if not customers:
        return None
    _assign_export_numbers(customers)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "客户汇总"
    summary_sheet.append(_SUMMARY_HEADERS)
    for row in _summary_rows(customers):
        summary_sheet.append(row)
    _style_sheet(
        summary_sheet,
        widths=[8, 24, 16, 12, 12, 12],
        amount_columns={3},
        number_columns={1, 4, 5, 6},
    )

    detail_sheet = workbook.create_sheet("明细清单")
    detail_sheet.append(_DETAIL_HEADERS)
    for row in _detail_rows(customers):
        detail_sheet.append(row)
    _style_sheet(
        detail_sheet,
        widths=[8, 14, 18, 12, 18, 10, 26, 20, 14, 14, 14, 14, 12, 12, 12, 18, 12, 14, 14, 14, 12],
        amount_columns={9, 10, 11, 12},
        unit_price_columns={13},
        number_columns={1, 6, 14, 16, 18, 19, 21},
        status_column=2,
        order_column=5,
    )
    _build_status_sheet(workbook)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_alerts_csv(db: Session, job_id: str) -> bytes:
    alerts = (
        db.query(Alert)
        .filter(Alert.job_id == job_id, Alert.status == "open")
        .order_by(Alert.created_at.desc())
        .all()
    )
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["alert_id", "job_id", "group_id", "alert_type", "severity", "status", "message", "created_at"])
    for a in alerts:
        writer.writerow([a.id, a.job_id, a.group_id, a.alert_type, a.severity, a.status, a.message, a.created_at.isoformat()])
    return output.getvalue().encode("utf-8-sig")



def export_customer_summary_csv(db: Session, job_id: str) -> bytes:
    rows = (
        db.query(Alert, MatchGroup)
        .join(MatchGroup, Alert.group_id == MatchGroup.id)
        .filter(Alert.job_id == job_id, Alert.status == "open")
        .all()
    )

    stats: dict[str, dict[str, Any]] = defaultdict(lambda: {"alert_count": 0, "high_count": 0, "medium_count": 0})
    for alert, group in rows:
        customer = ((group.summary_json or {}).get("aggregate") or {}).get("customer") or "未知客户"
        stats[customer]["alert_count"] += 1
        if alert.severity == "high":
            stats[customer]["high_count"] += 1
        else:
            stats[customer]["medium_count"] += 1

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["customer", "alert_count", "high_count", "medium_count"])
    for customer, item in sorted(stats.items(), key=lambda x: x[1]["alert_count"], reverse=True):
        writer.writerow([customer, item["alert_count"], item["high_count"], item["medium_count"]])
    return output.getvalue().encode("utf-8-sig")
