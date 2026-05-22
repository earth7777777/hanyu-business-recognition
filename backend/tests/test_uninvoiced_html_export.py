from __future__ import annotations

from io import BytesIO
from datetime import datetime, timezone

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.db.base import Base
from app.db.models import Alert, MatchGroup, NormalizedRecord, UploadJob, UploadedFile, ViewerCustomerAlertSetting
from app.services.export_service import export_uninvoiced_excel, export_uninvoiced_html


def _session():
    engine = create_engine(
        "sqlite+pysqlite://",
        future=True,
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)()


def _core(
    *,
    customer: str,
    order_no: str,
    line_no: str,
    item_code: str,
    item_name: str,
    quantity: float,
    amount: float | None,
    order_outbound_status: str,
    line_outbound_status: str,
    latest_outbound_date: str | None,
    executed_shipped_qty: float,
    invoiced_qty: float,
    uninvoiced_qty: float,
    biz_date: str = "2026-02-10",
    order_total_amount: float | None = None,
    tax_inclusive_unit_price: float | None = None,
) -> dict:
    return {
        "customer": customer,
        "customer_order_no": order_no,
        "entry_line_no": line_no,
        "item_code": item_code,
        "item_name": item_name,
        "quantity": quantity,
        "amount": amount,
        "order_total_amount": order_total_amount,
        "tax_inclusive_unit_price": (
            tax_inclusive_unit_price
            if tax_inclusive_unit_price is not None
            else (round(amount / quantity, 6) if amount is not None and quantity else None)
        ),
        "biz_date": biz_date,
        "order_outbound_status": order_outbound_status,
        "line_outbound_status": line_outbound_status,
        "latest_outbound_date": latest_outbound_date,
        "executed_shipped_qty": executed_shipped_qty,
        "invoiced_qty": invoiced_qty,
        "uninvoiced_qty": uninvoiced_qty,
        "order_unshipped_qty": max(quantity - executed_shipped_qty, 0),
    }


def _add_order_record(db, *, record_id: str, job_id: str, file_id: str, source_row: int, core: dict):
    db.add(
        NormalizedRecord(
            id=record_id,
            job_id=job_id,
            file_id=file_id,
            document_type="order",
            source_row=source_row,
            lifecycle_state="active",
            is_current_effective=True,
            payload_json={
                "core": core,
                "ext": {
                    "order_outbound_status_raw": "全部出库"
                    if core["order_outbound_status"] == "fully_outbound"
                    else "部分出库",
                    "line_outbound_status_raw": "全部出库"
                    if core["line_outbound_status"] == "fully_outbound"
                    else "部分出库",
                },
            },
        )
    )


def _add_alert(db, *, alert_id: str, job_id: str, group_id: str, record_id: str, core: dict, days: int):
    db.add(
        Alert(
            id=alert_id,
            job_id=job_id,
            group_id=group_id,
            alert_type="ship_after_no_finance",
            status="open",
            severity="medium",
            message=f"订单〔{core['customer_order_no']}〕距最近出库已〔{days}〕天，金额〔{core['amount']}〕，请尽快开票。",
            payload_json={
                "record_id": record_id,
                "source_row": 1,
                "customer_order_no": core["customer_order_no"],
                "entry_line_no": core["entry_line_no"],
                "item_code": core["item_code"],
                "item_name": core["item_name"],
                "quantity": core["quantity"],
                "amount": core["amount"],
                "tax_inclusive_unit_price": core["tax_inclusive_unit_price"],
                "biz_date": core["biz_date"],
                "order_outbound_status": core["order_outbound_status"],
                "line_outbound_status": core["line_outbound_status"],
                "latest_outbound_date": core["latest_outbound_date"],
                "days_after_outbound": days,
                "executed_shipped_qty": core["executed_shipped_qty"],
                "invoiced_qty": core["invoiced_qty"],
                "uninvoiced_qty": core["uninvoiced_qty"],
                "order_unshipped_qty": core["order_unshipped_qty"],
            },
        )
    )


def test_uninvoiced_html_groups_customer_order_and_product_context():
    db = _session()
    job = UploadJob(id="job-html", status="succeeded", created_by="test")
    file = UploadedFile(
        id="file-html",
        job_id=job.id,
        document_type="order",
        filename="order.csv",
        storage_path="/tmp/order.csv",
    )
    group = MatchGroup(
        id="group-html",
        job_id=job.id,
        group_key="group-html",
        summary_json={"aggregate": {"customer": "A客户"}},
    )
    db.add_all([job, file, group])

    should_core = _core(
        customer="A客户",
        order_no="SO-HTML-1",
        line_no="1",
        item_code="ITEM-A",
        item_name="产品A",
        quantity=100,
        amount=52000,
        order_outbound_status="partially_outbound",
        line_outbound_status="fully_outbound",
        latest_outbound_date="2026-03-01",
        executed_shipped_qty=100,
        invoiced_qty=20,
        uninvoiced_qty=80,
        order_total_amount=63000,
    )
    should_uninvoiced_core = _core(
        customer="A客户",
        order_no="SO-HTML-1",
        line_no="4",
        item_code="ITEM-D",
        item_name="产品D",
        quantity=5,
        amount=500,
        order_outbound_status="partially_outbound",
        line_outbound_status="fully_outbound",
        latest_outbound_date="2026-03-01",
        executed_shipped_qty=5,
        invoiced_qty=None,
        uninvoiced_qty=5,
        order_total_amount=63000,
    )
    done_core = _core(
        customer="A客户",
        order_no="SO-HTML-1",
        line_no="2",
        item_code="ITEM-B",
        item_name="产品B",
        quantity=10,
        amount=1000,
        order_outbound_status="partially_outbound",
        line_outbound_status="fully_outbound",
        latest_outbound_date="2026-03-01",
        executed_shipped_qty=10,
        invoiced_qty=10,
        uninvoiced_qty=0,
        order_total_amount=63000,
    )
    hold_core = _core(
        customer="A客户",
        order_no="SO-HTML-1",
        line_no="3",
        item_code="ITEM-C",
        item_name="产品C",
        quantity=20,
        amount=8000,
        order_outbound_status="partially_outbound",
        line_outbound_status="partially_outbound",
        latest_outbound_date="2026-03-01",
        executed_shipped_qty=5,
        invoiced_qty=0,
        uninvoiced_qty=20,
        order_total_amount=63000,
    )
    not_due_core = _core(
        customer="A客户",
        order_no="SO-HTML-1",
        line_no="5",
        item_code="ITEM-E",
        item_name="产品E",
        quantity=10,
        amount=1000,
        order_outbound_status="partially_outbound",
        line_outbound_status="fully_outbound",
        latest_outbound_date="2026-05-01",
        executed_shipped_qty=10,
        invoiced_qty=5,
        uninvoiced_qty=5,
        order_total_amount=63000,
    )
    no_alert_core = _core(
        customer="A客户",
        order_no="SO-NO-ALERT",
        line_no="1",
        item_code="ITEM-X",
        item_name="不应出现产品",
        quantity=8,
        amount=6000,
        order_outbound_status="partially_outbound",
        line_outbound_status="partially_outbound",
        latest_outbound_date="2026-03-01",
        executed_shipped_qty=2,
        invoiced_qty=0,
        uninvoiced_qty=8,
    )
    _add_order_record(db, record_id="rec-should", job_id=job.id, file_id=file.id, source_row=1, core=should_core)
    _add_order_record(db, record_id="rec-done", job_id=job.id, file_id=file.id, source_row=2, core=done_core)
    _add_order_record(db, record_id="rec-hold", job_id=job.id, file_id=file.id, source_row=3, core=hold_core)
    _add_order_record(db, record_id="rec-no-alert", job_id=job.id, file_id=file.id, source_row=4, core=no_alert_core)
    _add_order_record(
        db,
        record_id="rec-should-uninvoiced",
        job_id=job.id,
        file_id=file.id,
        source_row=5,
        core=should_uninvoiced_core,
    )
    _add_order_record(db, record_id="rec-not-due", job_id=job.id, file_id=file.id, source_row=6, core=not_due_core)
    _add_alert(db, alert_id="alert-html", job_id=job.id, group_id=group.id, record_id="rec-should", core=should_core, days=79)
    _add_alert(
        db,
        alert_id="alert-html-uninvoiced",
        job_id=job.id,
        group_id=group.id,
        record_id="rec-should-uninvoiced",
        core=should_uninvoiced_core,
        days=79,
    )
    db.commit()

    exported = export_uninvoiced_html(
        db,
        generated_at=datetime(2026, 5, 19, 13, 0, tzinfo=timezone.utc),
    )

    assert exported is not None
    html = exported.decode("utf-8")
    assert "超60天没开票" in html
    assert "A客户" in html
    assert "SO-HTML-1" in html
    assert "序号 1" in html
    assert "序号 2" in html
    assert "序号 3" not in html
    assert "产品A（ITEM-A）" in html
    assert "产品D（ITEM-D）" in html
    assert "产品B（ITEM-B）" not in html
    assert "产品C（ITEM-C）" not in html
    assert "产品E（ITEM-E）" not in html
    assert "已发完，应该催票" not in html
    assert "应该催票" not in html
    assert "产品已部分开票" in html
    assert "产品未开票" in html
    assert "开票状态暂无" not in html
    assert "已开完，不用催" not in html
    assert "还没发完，先不催" not in html
    assert "未发完暂不催金额" not in html
    assert "未发完暂不催金额合计" not in html
    assert "订单总金额" in html
    assert "¥63,000.00" in html
    assert "¥61,500.00" not in html
    assert "已开票金额" in html
    assert "¥11,900.00" in html
    assert "已开票金额</b>¥11,400.00" not in html
    assert "已开票金额</b>¥11,900.00（部分金额暂缺）" not in html
    assert "还差开票金额" in html
    assert "¥41,600.00" in html
    assert "¥42,100.00" in html
    assert "¥52,500.00" not in html
    assert "已开票金额</b>暂无" not in html
    assert "整单已发完 / 应催" not in html
    assert "部分发货 / 有产品应催" not in html
    assert "整单未发完" in html
    assert "客户 1/1" in html
    assert "订单 1/1" in html
    order_summary = html.split('<div class="order-summary">', 1)[1].split("</div>", 1)[0]
    assert order_summary.index("订单总金额") < order_summary.index("已开票金额")
    assert order_summary.index("已开票金额") < order_summary.index("还差开票金额")
    assert "订单总量" not in order_summary
    assert "已开票数量" not in order_summary
    assert "还差开票数量" not in order_summary
    assert ".order-summary { grid-template-columns: repeat(3, minmax(0, 1fr)); }" in html
    assert '<section class="report-section summary-report" aria-label="汇总">' in html
    assert '<section class="report-section detail-report" aria-label="明细">' in html
    assert '<div class="section-divider">汇总</div>' in html
    assert '<div class="section-divider">明细</div>' in html
    assert ".detail-report { break-before: page; page-break-before: always; }" in html
    assert "出库天数" in html
    assert "距最近出库" not in html
    assert "已超过60天" not in html
    assert "最长已超" not in html
    assert "SO-NO-ALERT" not in html
    assert "未解除" not in html
    assert "@media print" in html
    assert ".order-grid { display: grid; grid-template-columns: 1fr; gap: 12px; }" in html
    assert ".order-grid { grid-template-columns: 1fr; gap: 6mm; }" in html
    assert ".order-grid { grid-template-columns: repeat(2, minmax(0, 1fr));" not in html

    excel = export_uninvoiced_excel(
        db,
        generated_at=datetime(2026, 5, 19, 13, 0, tzinfo=timezone.utc),
    )
    assert excel is not None
    workbook = load_workbook(BytesIO(excel))
    assert workbook.sheetnames == ["客户汇总", "明细清单", "颜色说明"]
    summary = workbook["客户汇总"]
    detail = workbook["明细清单"]
    legend = workbook["颜色说明"]
    assert [cell.value for cell in detail[1]] == [
        "序号",
        "核对状态",
        "客户",
        "单据日期",
        "单据编号",
        "分录行号",
        "商品名称",
        "商品编码",
        "已开票金额",
        "未开票金额",
        "价税合计",
        "成交金额",
        "含税单价",
        "数量",
        "行出库状态",
        "行已执行已出库数量",
        "行开票状态",
        "行已开票数量",
        "行未开票数量",
        "最近出库日期",
        "已超天数",
    ]
    detail_rows = list(detail.iter_rows(min_row=2, values_only=True))
    assert len(detail_rows) == 5
    assert detail_rows[0] == (
        1,
        "应催",
        "A客户",
        "2026-02-10",
        "SO-HTML-1",
        "1",
        "产品A",
        "ITEM-A",
        10400,
        41600,
        52000,
        63000,
        520,
        100,
        "全部出库",
        100,
        "部分开票",
        20,
        80,
        "2026-03-01",
        19,
    )
    assert detail_rows[1] == (
        None,
        "已开完不催",
        "A客户",
        "2026-02-10",
        "SO-HTML-1",
        "2",
        "产品B",
        "ITEM-B",
        1000,
        None,
        1000,
        63000,
        100,
        10,
        "全部出库",
        10,
        "全部开票",
        10,
        None,
        "2026-03-01",
        19,
    )
    assert detail_rows[2] == (
        None,
        "未发完暂不催",
        "A客户",
        "2026-02-10",
        "SO-HTML-1",
        "3",
        "产品C",
        "ITEM-C",
        None,
        8000,
        8000,
        63000,
        400,
        20,
        "部分出库",
        5,
        "未开票",
        None,
        20,
        "2026-03-01",
        19,
    )
    assert detail_rows[3] == (
        2,
        "应催",
        "A客户",
        "2026-02-10",
        "SO-HTML-1",
        "4",
        "产品D",
        "ITEM-D",
        None,
        500,
        500,
        63000,
        100,
        5,
        "全部出库",
        5,
        "未开票",
        None,
        5,
        "2026-03-01",
        19,
    )
    assert detail_rows[4] == (
        None,
        "未满60天暂不催",
        "A客户",
        "2026-02-10",
        "SO-HTML-1",
        "5",
        "产品E",
        "ITEM-E",
        500,
        500,
        1000,
        63000,
        100,
        10,
        "全部出库",
        10,
        "部分开票",
        5,
        5,
        "2026-05-01",
        None,
    )
    assert detail.freeze_panes == "A2"
    assert detail.auto_filter.ref == "A1:U6"
    assert detail.sheet_view.showGridLines is True
    assert detail.print_options.gridLines is True
    assert detail["A1"].border.bottom.style == "medium"
    assert detail["A2"].border.left.style == "thin"
    assert detail["I2"].number_format == "#,##0.00"
    assert detail["J2"].number_format == "#,##0.00"
    assert detail["K2"].number_format == "#,##0.00"
    assert detail["L2"].number_format == "#,##0.00"
    assert detail["M2"].number_format == "#,##0.#####"
    assert detail["N2"].number_format == "#,##0"
    assert detail["P2"].number_format == "#,##0"
    assert detail["R2"].number_format == "#,##0"
    assert detail["S2"].number_format == "#,##0"
    assert detail["U2"].number_format == "#,##0"
    assert detail["I2"].value == 10400
    assert detail["J2"].value == 41600
    assert detail["K2"].value == 52000
    assert detail["L2"].value == 63000
    assert detail["M2"].value == 520
    assert detail["N2"].value == 100
    assert detail["P2"].value == 100
    assert detail["R2"].value == 20
    assert detail["S2"].value == 80
    assert [cell.value for cell in summary[1]] == ["序号", "客户", "应催金额", "应催订单数", "应催产品行数", "最长出库天数"]
    assert [cell.value for cell in summary[2]] == [1, "A客户", 42100, 1, 2, 79]
    assert [cell.value for cell in legend[1]] == ["颜色", "核对状态", "说明"]
    assert [cell.value for cell in legend[2]][1:] == [
        "应催",
        "这条产品行已经出库满60天，并且还有未开票数量，需要催票。",
    ]
    assert [cell.value for cell in legend[4]][1:] == ["已开完不催", "这条产品行已经开票完成，不需要催票。"]
    assert legend["A2"].fill.fgColor.rgb == "00FCE4D6"


def test_uninvoiced_html_exposes_missing_source_amounts():
    db = _session()
    job = UploadJob(id="job-missing-amount", status="succeeded", created_by="test")
    file = UploadedFile(
        id="file-missing-amount",
        job_id=job.id,
        document_type="order",
        filename="order.csv",
        storage_path="/tmp/order.csv",
    )
    group = MatchGroup(
        id="group-missing-amount",
        job_id=job.id,
        group_key="group-missing-amount",
        summary_json={"aggregate": {"customer": "金额缺失客户"}},
    )
    db.add_all([job, file, group])

    core = _core(
        customer="金额缺失客户",
        order_no="SO-MISSING-AMOUNT",
        line_no="1",
        item_code="ITEM-MISSING",
        item_name="金额缺失产品",
        quantity=15,
        amount=0,
        order_outbound_status="fully_outbound",
        line_outbound_status="fully_outbound",
        latest_outbound_date="2026-01-10",
        executed_shipped_qty=15,
        invoiced_qty=0,
        uninvoiced_qty=15,
        tax_inclusive_unit_price=0,
    )
    _add_order_record(db, record_id="rec-missing-amount", job_id=job.id, file_id=file.id, source_row=1, core=core)
    _add_alert(
        db,
        alert_id="alert-missing-amount",
        job_id=job.id,
        group_id=group.id,
        record_id="rec-missing-amount",
        core=core,
        days=131,
    )
    db.commit()

    exported = export_uninvoiced_html(
        db,
        generated_at=datetime(2026, 5, 21, 18, 30, tzinfo=timezone.utc),
    )
    assert exported is not None
    html = exported.decode("utf-8")
    order_block = html.split("<span>SO-MISSING-AMOUNT</span>", 1)[1].split("</article>", 1)[0]
    assert "订单总金额</b>金额暂缺" in order_block
    assert "已开票金额</b>金额暂缺" in order_block
    assert "还差开票金额</b>金额暂缺" in order_block
    assert "<strong>金额暂缺</strong>" in order_block
    assert "订单总金额</b>¥0.00" not in order_block
    assert "已开票金额</b>¥0.00" not in order_block
    assert "还差开票金额</b>¥0.00" not in order_block

    excel = export_uninvoiced_excel(
        db,
        generated_at=datetime(2026, 5, 21, 18, 30, tzinfo=timezone.utc),
    )
    assert excel is not None
    workbook = load_workbook(BytesIO(excel))
    detail = workbook["明细清单"]
    assert detail["J2"].value is None
    assert detail["K2"].value is None
    assert detail["M2"].value is None


def test_uninvoiced_export_keeps_keyword_customers_adjacent_without_merging():
    db = _session()
    job = UploadJob(id="job-customer-sort", status="succeeded", created_by="test")
    file = UploadedFile(
        id="file-customer-sort",
        job_id=job.id,
        document_type="order",
        filename="order.csv",
        storage_path="/tmp/order.csv",
    )
    db.add_all([job, file])

    rows = [
        ("杭州巨星工具", "HZ-JX-1", "巨星产品A", "JX-A", 1000),
        ("安吉热威", "AJ-RW-1", "热威产品A", "RW-A", 900),
        ("浙江巨星工具", "ZJ-JX-1", "巨星产品B", "JX-B", 100),
        ("杭州热威", "HZ-RW-1", "热威产品B", "RW-B", 50),
        ("杭州巨星工具", "HZ-JX-2", "巨星产品C", "JX-C", 200),
    ]
    for index, (customer, order_no, item_name, item_code, amount) in enumerate(rows, start=1):
        group = MatchGroup(
            id=f"group-customer-sort-{index}",
            job_id=job.id,
            group_key=f"group-customer-sort-{index}",
            summary_json={"aggregate": {"customer": customer}},
        )
        core = _core(
            customer=customer,
            order_no=order_no,
            line_no="1",
            item_code=item_code,
            item_name=item_name,
            quantity=10,
            amount=amount,
            order_outbound_status="fully_outbound",
            line_outbound_status="fully_outbound",
            latest_outbound_date="2026-03-01",
            executed_shipped_qty=10,
            invoiced_qty=0,
            uninvoiced_qty=10,
        )
        record_id = f"rec-customer-sort-{index}"
        db.add(group)
        _add_order_record(db, record_id=record_id, job_id=job.id, file_id=file.id, source_row=index, core=core)
        _add_alert(
            db,
            alert_id=f"alert-customer-sort-{index}",
            job_id=job.id,
            group_id=group.id,
            record_id=record_id,
            core=core,
            days=79,
        )
    db.commit()

    exported = export_uninvoiced_html(
        db,
        generated_at=datetime(2026, 5, 19, 13, 0, tzinfo=timezone.utc),
    )
    assert exported is not None
    html = exported.decode("utf-8")
    assert html.count("<h2>杭州巨星工具</h2>") == 2
    assert html.count("<h2>浙江巨星工具</h2>") == 2
    assert html.count("<h2>安吉热威</h2>") == 2
    assert html.count("<h2>杭州热威</h2>") == 2
    assert html.index("<h2>杭州巨星工具</h2>") < html.index("<h2>浙江巨星工具</h2>")
    assert html.index("<h2>浙江巨星工具</h2>") < html.index("<h2>安吉热威</h2>")
    assert html.index("<h2>安吉热威</h2>") < html.index("<h2>杭州热威</h2>")
    hangzhou_juxing_section = html.split("<h2>杭州巨星工具</h2>", 1)[1].split("</section>", 1)[0]
    assert "相关订单数</b>2 笔" in hangzhou_juxing_section
    assert "应催金额</b>¥1,200.00" in hangzhou_juxing_section

    excel = export_uninvoiced_excel(
        db,
        generated_at=datetime(2026, 5, 19, 13, 0, tzinfo=timezone.utc),
    )
    assert excel is not None
    workbook = load_workbook(BytesIO(excel))
    detail_rows = list(workbook["明细清单"].iter_rows(min_row=2, values_only=True))
    customers = [row[2] for row in detail_rows]
    collapsed_customers: list[str] = []
    for customer in customers:
        if not collapsed_customers or customer != collapsed_customers[-1]:
            collapsed_customers.append(customer)
    assert collapsed_customers == ["杭州巨星工具", "浙江巨星工具", "安吉热威", "杭州热威"]


def test_uninvoiced_export_sorts_customer_orders_by_outbound_then_order_date():
    db = _session()
    job = UploadJob(id="job-order-sort", status="succeeded", created_by="test")
    file = UploadedFile(
        id="file-order-sort",
        job_id=job.id,
        document_type="order",
        filename="order.csv",
        storage_path="/tmp/order.csv",
    )
    db.add_all([job, file])

    rows = [
        {
            "order_no": "SO-FULL-OLD-HIGH",
            "biz_date": "2026-01-10",
            "amount": 9000,
            "order_outbound_status": "fully_outbound",
            "line_outbound_status": "fully_outbound",
        },
        {
            "order_no": "SO-FULL-NEW-LOW",
            "biz_date": "2026-03-20",
            "amount": 100,
            "order_outbound_status": "fully_outbound",
            "line_outbound_status": "fully_outbound",
        },
        {
            "order_no": "SO-PARTIAL-NEW-HIGH",
            "biz_date": "2026-04-20",
            "amount": 12000,
            "order_outbound_status": "partially_outbound",
            "line_outbound_status": "fully_outbound",
        },
    ]
    for index, row in enumerate(rows, start=1):
        group = MatchGroup(
            id=f"group-order-sort-{index}",
            job_id=job.id,
            group_key=f"group-order-sort-{index}",
            summary_json={"aggregate": {"customer": "排序客户"}},
        )
        core = _core(
            customer="排序客户",
            order_no=row["order_no"],
            line_no="1",
            item_code=f"ITEM-{index}",
            item_name=f"产品{index}",
            quantity=10,
            amount=row["amount"],
            order_outbound_status=row["order_outbound_status"],
            line_outbound_status=row["line_outbound_status"],
            latest_outbound_date="2026-02-01",
            executed_shipped_qty=10,
            invoiced_qty=0,
            uninvoiced_qty=10,
            biz_date=row["biz_date"],
        )
        record_id = f"rec-order-sort-{index}"
        db.add(group)
        _add_order_record(db, record_id=record_id, job_id=job.id, file_id=file.id, source_row=index, core=core)
        _add_alert(
            db,
            alert_id=f"alert-order-sort-{index}",
            job_id=job.id,
            group_id=group.id,
            record_id=record_id,
            core=core,
            days=90,
        )
    db.commit()

    exported = export_uninvoiced_html(
        db,
        generated_at=datetime(2026, 5, 21, 8, 0, tzinfo=timezone.utc),
    )
    assert exported is not None
    html = exported.decode("utf-8")
    assert html.index("SO-FULL-NEW-LOW") < html.index("SO-FULL-OLD-HIGH")
    assert html.index("SO-FULL-OLD-HIGH") < html.index("SO-PARTIAL-NEW-HIGH")

    excel = export_uninvoiced_excel(
        db,
        generated_at=datetime(2026, 5, 21, 8, 0, tzinfo=timezone.utc),
    )
    assert excel is not None
    workbook = load_workbook(BytesIO(excel))
    detail_orders = [row[4] for row in workbook["明细清单"].iter_rows(min_row=2, values_only=True)]
    assert detail_orders == ["SO-FULL-NEW-LOW", "SO-FULL-OLD-HIGH", "SO-PARTIAL-NEW-HIGH"]


def test_uninvoiced_html_skips_disabled_customer():
    db = _session()
    job = UploadJob(id="job-disabled", status="succeeded", created_by="test")
    file = UploadedFile(
        id="file-disabled",
        job_id=job.id,
        document_type="order",
        filename="order.csv",
        storage_path="/tmp/order.csv",
    )
    group = MatchGroup(
        id="group-disabled",
        job_id=job.id,
        group_key="group-disabled",
        summary_json={"aggregate": {"customer": "B客户"}},
    )
    db.add_all([job, file, group])
    core = _core(
        customer="B客户",
        order_no="SO-DISABLED",
        line_no="1",
        item_code="ITEM-D",
        item_name="产品D",
        quantity=10,
        amount=12000,
        order_outbound_status="fully_outbound",
        line_outbound_status="fully_outbound",
        latest_outbound_date="2026-03-01",
        executed_shipped_qty=10,
        invoiced_qty=0,
        uninvoiced_qty=10,
    )
    _add_order_record(db, record_id="rec-disabled", job_id=job.id, file_id=file.id, source_row=1, core=core)
    _add_alert(db, alert_id="alert-disabled", job_id=job.id, group_id=group.id, record_id="rec-disabled", core=core, days=79)
    db.add(
        ViewerCustomerAlertSetting(
            id="setting-disabled",
            customer_name="B客户",
            customer_key="b客户",
            alert_type="ship_after_no_finance",
            is_enabled=False,
            last_reason="允许延期",
            last_operator_name="月总",
        )
    )
    db.commit()

    assert export_uninvoiced_html(db) is None
    assert export_uninvoiced_excel(db) is None
